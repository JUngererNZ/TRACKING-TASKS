import json
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def generate_enroute_site_mapping(file_path, output_json_path):
    if not os.path.exists(file_path):
        print(f"Error: The file '{file_path}' does not exist.")
        return

    print(f"Loading workbook: {file_path}...")
    # Using read_only and data_only for quick and memory-efficient reading
    wb = load_workbook(file_path, read_only=True, data_only=True)
    
    target_sheet = "ENROUTE SITE"
    if target_sheet not in wb.sheetnames:
        print(f"Error: Sheet '{target_sheet}' not found in the workbook.")
        print(f"Available sheets: {wb.sheetnames}")
        return

    ws = wb[target_sheet]
    print(f"Mapping sheet: {target_sheet}...")

    # Define exact rows based on structure: row 7 is headers, row 8 is data
    HEADER_ROW_NUM = 7
    DATA_ROW_NUM = 8

    headers = []
    sample_values = []

    # Extract row 7 and row 8 cleanly
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == HEADER_ROW_NUM:
            headers = list(row)
        elif row_idx == DATA_ROW_NUM:
            sample_values = list(row)
            break  # We have everything we need, exit loop early

    # Safeguard in case row 8 is shorter than row 7
    while len(sample_values) < len(headers):
        sample_values.append(None)

    sheet_columns = []

    # Build the column mapping
    for col_idx, header_value in enumerate(headers, start=1):
        # Skip completely blank columns if they trailing off the end of the sheet
        if header_value is None and (col_idx > len(sample_values) or sample_values[col_idx-1] is None):
            # If both header and sample are None, check if there are more columns left
            if all(h is None for h in headers[col_idx-1:]):
                break

        clean_header = str(header_value).strip() if header_value is not None else f"Unnamed_Column_{col_idx}"
        column_letter = get_column_letter(col_idx)
        raw_sample = sample_values[col_idx - 1]
        
        # Format dates/numbers nicely for the JSON string representation if needed
        sample_string = str(raw_sample).strip() if raw_sample is not None else ""

        sheet_columns.append({
            "column_index": col_idx,
            "column_letter": column_letter,
            "header_name": clean_header,
            "sample_row_8_value": sample_string
        })

    # Prepare final JSON layout
    mapping_output = {
        "workbook": os.path.basename(file_path),
        "sheet": target_sheet,
        "header_row": HEADER_ROW_NUM,
        "first_data_row": DATA_ROW_NUM,
        "columns": sheet_columns
    }

    # Write mapping dictionary to JSON
    with open(output_json_path, "w", encoding="utf-8") as f:
        json.dump(mapping_output, f, indent=4, ensure_ascii=False)
        
    print(f"\nSuccess! Mapping written to: {output_json_path}")

if __name__ == "__main__":
    EXCEL_FILE = r"C:\Users\Jason\Projects\TRACKING-TASKS\company BARTRAC - KAMOA TRACKING AS OF 03-06-2026.xlsx"
    OUTPUT_JSON = r"C:\Users\Jason\Projects\TRACKING-TASKS\JSON-REF\bartrac-kamoa-enroute-site-mapping.json"
    
    generate_enroute_site_mapping(EXCEL_FILE, OUTPUT_JSON)