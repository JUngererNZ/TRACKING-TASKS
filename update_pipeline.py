import os
import json
import openpyxl
import re

def normalize_key(value):
    """
    Strips spaces, trailing tabs (\t), dashes, and non-alphanumeric text.
    Ensures 'FV44CKGP\t', 'FV 44 CK GP' and 'fv44ckgp' match flawlessly. 15.06.26 11:24
    """
    if not value:
        return ""
    return re.sub(r'[^A-Z0-9]', '', str(value).upper())

def load_json_mapping(mapping_dir, workbook_name):
    """Loads JSON configurations using flexible client token pairs."""
    name_upper = workbook_name.upper()
    
    for file in os.listdir(mapping_dir):
        if not file.endswith(".json"):
            continue
            
        file_upper = file.upper()
        target_path = os.path.join(mapping_dir, file)
        
        is_match = False
        if "FML" in name_upper and "CAT" in name_upper and "FML" in file_upper:
            is_match = True
        else:
            for token in ["CONGO", "ERG", "KAMOA", "KCC", "MUMI", "BONDED", "KANU"]:
                if token in name_upper and token in file_upper:
                    is_match = True
                    break
                    
        if is_match:
            try:
                if os.path.getsize(target_path) == 0:
                    return None
                with open(target_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except json.JSONDecodeError:
                return None
                
    return None

def extract_vendor_updates(vendor_file_path, mapping_guide):
    """
    Stages updates from the horizontal matrix format of FML CAT 6060.
    Finds the 'Truck' label row automatically, then reads the latest row data status below it.
    """
    wb = openpyxl.load_workbook(vendor_file_path, data_only=True)
    sheet_name = mapping_guide["sheet"] if (mapping_guide and "sheet" in mapping_guide) else "Sheet1"
    
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
        
    ws = wb[sheet_name]
    vendor_updates = {}

    # Scan to find the row index containing 'Truck' or registration signatures dynamically
    truck_row = None
    for r_idx in range(1, 15):
        cell_check = str(ws.cell(row=r_idx, column=1).value or "").strip().upper()
        if "TRUCK" in cell_check or "FLEET" in cell_check:
            truck_row = r_idx
            break
            
    if not truck_row:
        truck_row = 10  # Standard matrix layout row fallback

    latest_data_row = ws.max_row
    
    for col_idx in range(2, ws.max_column + 1):
        truck_val = ws.cell(row=truck_row, column=col_idx).value
        if not truck_val:
            continue
            
        status_val = None
        # Back-track up the column to get the most recent daily log description update entry
        for r_idx in range(latest_data_row, truck_row, -1):
            cell_check = ws.cell(row=r_idx, column=col_idx).value
            if cell_check:
                status_val = str(cell_check).strip()
                break
                
        if truck_val and status_val:
            clean_truck = normalize_key(truck_val)
            vendor_updates[clean_truck] = status_val
            
    print(f"Successfully staged {len(vendor_updates)} asset tracking statuses from vendor report.")
    return vendor_updates

def process_all_company_updates(company_files_dir, vendor_file_path, mapping_dir):
    """Iterates through tracking workbooks and coordinates the write-back engine."""
    if not os.path.exists(vendor_file_path):
        print(f"Error: Master vendor file not found at: {vendor_file_path}")
        return

    vendor_file_name = os.path.basename(vendor_file_path)
    vendor_schema = load_json_mapping(mapping_dir, vendor_file_name)
    vendor_updates = extract_vendor_updates(vendor_file_path, vendor_schema)
    
    if not vendor_updates:
        print("No vendor update payload found. Aborting execution.")
        return

    for filename in os.listdir(company_files_dir):
        if not filename.endswith(".xlsx") or filename.startswith("~$") or filename == vendor_file_name:
            continue
            
        print(f"\n" + "="*50)
        print(f"Beginning update sequence for: {filename}")
        
        schema_guide = load_json_mapping(mapping_dir, filename)
        execute_mapped_update(
            file_path=os.path.join(company_files_dir, filename),
            mapping_guide=schema_guide,
            updates_lookup=vendor_updates
        )

def execute_mapped_update(file_path, mapping_guide, updates_lookup):
    """Dynamically detects column indexes via header text mapping matching rules."""
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    sheet_name = None
    if mapping_guide and "sheet" in mapping_guide:
        sheet_name = mapping_guide["sheet"]
    else:
        for potential_name in ["ENROUTE SITE", "CURRENT SHIPMENTS ", "CURRENT SHIPMENTS"]:
            if potential_name in wb.sheetnames:
                sheet_name = potential_name
                break
                
    if not sheet_name or sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
        
    ws = wb[sheet_name]
    
    truck_idx = None
    status_idx = None
    
    # DYNAMIC SEARCH 1: Find columns based on header labels across the first 15 rows
    for r_idx in range(1, 15):
        for col_idx in range(1, ws.max_column + 1):
            cell_val = str(ws.cell(row=r_idx, column=col_idx).value or "").strip().upper()
            if "HORSE REG" in cell_val or ("TRUCK" in cell_val and not truck_idx):
                truck_idx = col_idx
            elif "ACTUAL STATUS" in cell_val or ("STATUS" in cell_val and not status_idx):
                status_idx = col_idx
        if truck_idx and status_idx:
            break

    # Hardcoded fallbacks if headers cannot be found textually
    if not truck_idx:
        truck_idx = 12
    if not status_idx:
        status_idx = 15

    updated_count = 0

    # DYNAMIC SEARCH 2: Scan ALL rows from top to bottom to protect against layout drift
    for row_idx in range(1, ws.max_row + 1):
        truck_val = ws.cell(row=row_idx, column=truck_idx).value
        if not truck_val:
            continue
            
        clean_truck = normalize_key(truck_val)
        
        # Avoid matching row headers text themselves
        if "HORSE" in clean_truck or "REG" in clean_truck or "TRUCK" in clean_truck:
            continue
            
        if clean_truck in updates_lookup:
            ws.cell(row=row_idx, column=status_idx).value = updates_lookup[clean_truck]
            updated_count += 1
            
    if updated_count > 0:
        wb.save(file_path)
        print(f"Successfully updated {updated_count} rows on sheet '{sheet_name}'.")
    else:
        print(f"No matching truck records found for updates on sheet '{sheet_name}'. (Scanned Truck Col: {truck_idx}, Status Col: {status_idx})")

if __name__ == "__main__":
    COMPANY_DIR = r"C:\Users\Jason\Projects\TRACKING-TASKS\15-06-2026"
    MAPPING_DIR = r"C:\Users\Jason\Projects\TRACKING-TASKS\JSON-REF"          
    VENDOR_FILE_PATH = r"C:\Users\Jason\Projects\TRACKING-TASKS\vendor_reports\FML CAT 6060.xlsx"
    
    process_all_company_updates(
        company_files_dir=COMPANY_DIR,
        vendor_file_path=VENDOR_FILE_PATH,
        mapping_dir=MAPPING_DIR
    )